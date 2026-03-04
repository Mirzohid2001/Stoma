from datetime import datetime
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..models import Client, Expense, Order, Payment

# Excel umumiy stillar
_EXCEL_HEADER_FILL = PatternFill(start_color='0d9488', end_color='0d9488', fill_type='solid')
_EXCEL_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
_EXCEL_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
_EXCEL_TOTAL_FILL = PatternFill(start_color='E0F2F1', end_color='E0F2F1', fill_type='solid')
_EXCEL_TOTAL_FONT = Font(bold=True)
_EXCEL_ALT_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')


def _style_excel_sheet(ws, num_cols, currency_cols=None, total_row=None, column_widths=None):
    """Excel varaqiga sarlavha, chegaralar va ustun enlarini qo'llash."""
    currency_cols = currency_cols or []
    column_widths = column_widths or []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=num_cols):
        for cell in row:
            cell.border = _EXCEL_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for c in range(1, num_cols + 1):
        ws.cell(row=1, column=c).fill = _EXCEL_HEADER_FILL
        ws.cell(row=1, column=c).font = _EXCEL_HEADER_FONT
    ws.row_dimensions[1].height = 22
    data_end = (total_row - 1) if total_row is not None else ws.max_row
    for row in range(2, data_end + 1):
        if (row - 2) % 2 == 1:
            for c in range(1, num_cols + 1):
                ws.cell(row=row, column=c).fill = _EXCEL_ALT_FILL
    for col_idx in currency_cols:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
    if total_row is not None:
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=total_row, column=c)
            cell.fill = _EXCEL_TOTAL_FILL
            cell.font = _EXCEL_TOTAL_FONT
        for col_idx in currency_cols:
            ws.cell(row=total_row, column=col_idx).alignment = Alignment(horizontal='right', vertical='center')
    for i, w in enumerate(column_widths):
        if i < num_cols:
            ws.column_dimensions[get_column_letter(i + 1)].width = w
    if ws.max_row >= 2:
        ws.freeze_panes = 'A2'


def _parse_date(s):
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


@login_required
def export_clients_excel(request):
    clients = Client.objects.all().order_by('full_name')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mijozlar'
    ws.append(['To\'liq ism', 'Telefon', 'Buyurtmalar soni', 'Jami to\'langan (so\'m)', 'Qarz (so\'m)', 'Qo\'shilgan sana'])
    for c in clients:
        ws.append([
            c.full_name,
            c.phone or '',
            c.orders_count,
            float(c.total_spent),
            float(c.total_debt),
            c.created_at.strftime('%Y-%m-%d') if c.created_at else '',
        ])
    _style_excel_sheet(ws, num_cols=6, currency_cols=[4, 5], column_widths=[28, 18, 14, 20, 18, 14])
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename=mijozlar_{timezone.now().date()}.xlsx'
    wb.save(resp)
    return resp


@login_required
def export_expenses_excel(request):
    from datetime import timedelta
    today = timezone.now().date()
    from_date = _parse_date(request.GET.get('from')) or (today.replace(day=1))
    to_date = _parse_date(request.GET.get('to')) or today

    expenses = Expense.objects.filter(
        expense_date__gte=from_date,
        expense_date__lte=to_date
    ).order_by('-expense_date')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rasxodlar'
    ws.append(['Sana', 'Turi', 'Summa (so\'m)', 'Tavsif'])
    for e in expenses:
        ws.append([str(e.expense_date), e.get_category_display(), float(e.amount), e.description or ''])
    total = expenses.aggregate(s=Sum('amount'))['s'] or 0
    total_row = ws.max_row + 1
    ws.append(['', '', float(total), 'Jami'])
    _style_excel_sheet(ws, num_cols=4, currency_cols=[3], total_row=total_row, column_widths=[14, 20, 18, 36])
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename=rasxodlar_{from_date}_{to_date}.xlsx'
    wb.save(resp)
    return resp


@login_required
def export_sales_excel(request):
    from_date = _parse_date(request.GET.get('from')) or (timezone.now().date() - __import__('datetime').timedelta(days=30))
    to_date = _parse_date(request.GET.get('to')) or timezone.now().date()

    payments = Payment.objects.filter(
        payment_date__gte=from_date,
        payment_date__lte=to_date
    ).select_related('order', 'order__client').order_by('payment_date')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Savdo hisoboti'
    ws.append(['Sana', 'Mijoz', 'Buyurtma', 'Summa (so\'m)'])
    for p in payments:
        ws.append([str(p.payment_date), p.order.client.full_name, p.order.order_number, float(p.amount)])
    total = payments.aggregate(s=Sum('amount'))['s'] or 0
    total_row = ws.max_row + 1
    ws.append(['', '', 'Jami', float(total)])
    _style_excel_sheet(ws, num_cols=4, currency_cols=[4], total_row=total_row, column_widths=[14, 32, 16, 18])
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename=savdo_{from_date}_{to_date}.xlsx'
    wb.save(resp)
    return resp


@login_required
def export_sales_pdf(request):
    from_date = _parse_date(request.GET.get('from')) or (timezone.now().date() - __import__('datetime').timedelta(days=30))
    to_date = _parse_date(request.GET.get('to')) or timezone.now().date()

    payments = list(Payment.objects.filter(
        payment_date__gte=from_date,
        payment_date__lte=to_date
    ).select_related('order', 'order__client').order_by('payment_date'))

    total = sum(float(p.amount) for p in payments)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph('Savdo hisoboti', styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f'Sana: {from_date} — {to_date}', styles['Normal']))
    elements.append(Spacer(1, 20))

    data = [['Sana', 'Mijoz', 'Buyurtma', 'Summa']]
    for p in payments[:100]:
        data.append([str(p.payment_date), p.order.client.full_name[:30], p.order.order_number, f'{float(p.amount):,.0f}'])
    if len(payments) > 100:
        data.append(['...', f'{len(payments)-100} ta yana', '', ''])
    data.append(['', '', 'Jami', f'{total:,.0f} so\'m'])

    t = Table(data, colWidths=[60, 120, 80, 80])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d9488')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
    ]))
    elements.append(t)
    doc.build(elements)

    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename=savdo_{from_date}_{to_date}.pdf'
    return resp


@login_required
def export_debts_excel(request):
    today = timezone.now().date()
    debtors = []
    seen = set()
    for order in Order.objects.filter(status__in=['draft', 'in_progress']).select_related('client'):
        if order.remaining_debt <= 0 or order.client_id in seen:
            continue
        seen.add(order.client_id)
        dl = order.debt_payment_deadline
        debtors.append({
            'client': order.client,
            'debt': order.client.total_debt,
            'deadline': dl,
            'days_left': (dl - today).days if dl else None,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = 'Qarzdorlar'
    ws.append(['Mijoz', 'Telefon', 'Qarz (so\'m)', 'To\'lov sanasi', 'Qolgan kun'])
    for d in debtors:
        ws.append([
            d['client'].full_name,
            d['client'].phone or '',
            float(d['debt']),
            str(d['deadline']) if d['deadline'] else '',
            d['days_left'] if d['days_left'] is not None else '',
        ])
    _style_excel_sheet(ws, num_cols=5, currency_cols=[3], column_widths=[28, 18, 18, 14, 12])
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename=qarzdorlar.xlsx'
    wb.save(resp)
    return resp


@login_required
def export_debts_pdf(request):
    today = timezone.now().date()
    debtors = []
    seen = set()
    for order in Order.objects.filter(status__in=['draft', 'in_progress']).select_related('client'):
        if order.remaining_debt <= 0 or order.client_id in seen:
            continue
        seen.add(order.client_id)
        dl = order.debt_payment_deadline
        days_left = (dl - today).days if dl else None
        debtors.append({
            'client': order.client,
            'debt': float(order.client.total_debt),
            'deadline': str(dl) if dl else '—',
            'days_left': str(days_left) if days_left is not None else '—',
        })

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph('Qarzdorlar hisoboti', styles['Title']))
    elements.append(Spacer(1, 20))

    data = [['Mijoz', 'Telefon', 'Qarz', "To'lov sanasi", 'Kun']]
    for d in debtors:
        data.append([d['client'].full_name[:25], d['client'].phone[:15], f"{d['debt']:,.0f}", d['deadline'], d['days_left']])

    t = Table(data, colWidths=[100, 80, 80, 80, 50])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d9488')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(t)
    doc.build(elements)

    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename=qarzdorlar.pdf'
    return resp
